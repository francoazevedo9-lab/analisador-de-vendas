import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Vendas"

# Cores
azul = "1565C0"
azul_claro = "E3F2FD"
branco = "FFFFFF"
cinza = "F5F5F5"

# Tamanho das colunas
ws.column_dimensions['A'].width = 16
ws.column_dimensions['B'].width = 22
ws.column_dimensions['C'].width = 14
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 16

# Título
ws.merge_cells('A1:E1')
ws['A1'] = '📊 CONTROLE DE VENDAS'
ws['A1'].font = Font(name='Calibri', size=16, bold=True, color=branco)
ws['A1'].fill = PatternFill("solid", fgColor=azul)
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 40

# Instrução
ws.merge_cells('A2:E2')
ws['A2'] = 'Preencha os dados abaixo e execute o script gerar_relatorio.py para gerar o PDF'
ws['A2'].font = Font(name='Calibri', size=10, italic=True, color="546E7A")
ws['A2'].fill = PatternFill("solid", fgColor=azul_claro)
ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[2].height = 24

# Linha vazia
ws.row_dimensions[3].height = 8

# Cabeçalhos
headers = ['Data', 'Produto', 'Quantidade', 'Preço Unitário', 'Total']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col)
    cell.value = header
    cell.font = Font(name='Calibri', size=11, bold=True, color=branco)
    cell.fill = PatternFill("solid", fgColor=azul)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[4].height = 32

# Borda
thin = Side(style='thin', color="BBDEFB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Dados de exemplo + fórmulas
dados_exemplo = [
    ('2024-01-03', 'Camiseta', 3, 49.90),
    ('2024-01-03', 'Calça', 1, 89.90),
    ('2024-01-05', 'Camiseta', 2, 49.90),
    ('2024-01-07', 'Tênis', 1, 199.90),
    ('2024-01-10', 'Calça', 2, 89.90),
]

for i, (data, produto, qtd, preco) in enumerate(dados_exemplo):
    row = 5 + i
    bg = azul_claro if i % 2 == 0 else branco
    
    ws.cell(row=row, column=1).value = data
    ws.cell(row=row, column=2).value = produto
    ws.cell(row=row, column=3).value = qtd
    ws.cell(row=row, column=4).value = preco
    ws.cell(row=row, column=5).value = f'=C{row}*D{row}'
    
    for col in range(1, 6):
        cell = ws.cell(row=row, column=col)
        cell.fill = PatternFill("solid", fgColor=bg.replace("#",""))
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(name='Calibri', size=10)
    
    ws.row_dimensions[row].height = 24

# Linhas vazias para o cliente preencher
for i in range(10):
    row = 10 + i
    bg = azul_claro if i % 2 == 0 else branco
    ws.cell(row=row, column=5).value = f'=IF(C{row}="","",C{row}*D{row})'
    for col in range(1, 6):
        cell = ws.cell(row=row, column=col)
        cell.fill = PatternFill("solid", fgColor=bg.replace("#",""))
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(name='Calibri', size=10)
    ws.row_dimensions[row].height = 24

# Total geral
row_total = 21
ws.merge_cells(f'A{row_total}:D{row_total}')
ws[f'A{row_total}'] = 'TOTAL GERAL'
ws[f'A{row_total}'].font = Font(name='Calibri', size=11, bold=True, color=branco)
ws[f'A{row_total}'].fill = PatternFill("solid", fgColor=azul)
ws[f'A{row_total}'].alignment = Alignment(horizontal='center', vertical='center')
ws[f'E{row_total}'] = f'=SUM(E5:E20)'
ws[f'E{row_total}'].font = Font(name='Calibri', size=11, bold=True, color=branco)
ws[f'E{row_total}'].fill = PatternFill("solid", fgColor=azul)
ws[f'E{row_total}'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[row_total].height = 32

wb.save('planilha_vendas.xlsx')
print("✅ Planilha criada: planilha_vendas.xlsx")